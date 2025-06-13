import json
import pandas as pd
import requests
from pathlib import Path
import re
from concurrent.futures import ThreadPoolExecutor
import time
import os
from openpyxl.styles import PatternFill, Font

# API configuration - should be loaded from environment or config in production
API_KEY = "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"  # Replace with your API key or load from environment
API_URL = "https://openrouter.ai/api/v1/chat/completions"

def load_compliance_results():
    """Load the compliance analysis results from Excel file."""
    base_dir = Path(r"C:\Users\91810\OneDrive\Desktop\CPC_AIB_Life")
    excel_file_path = base_dir / "output" / "enhanced_document_analysis" / "compliance_analysis_report.xlsx"
    
    print(f"Loading compliance results from {excel_file_path}...")
    df = pd.read_excel(excel_file_path)
    
    # Convert Excel data to the format needed for processing
    # Group by document_name to reconstruct the original JSON structure
    grouped_data = []
    for document_name, group in df.groupby('document_name'):
        detailed_results = []
        for _, row in group.iterrows():
            # Convert all values to strings to avoid type issues
            result = {
                'section_type': str(row.get('section_type', '')),
                'source_name': str(row.get('source_name', '')),
                'regulation_number': str(row.get('regulation_number', '')),
                'regulation_title': str(row.get('regulation_title', '')),
                'regulation_text': str(row.get('regulation_text', '')),
                'is_compliant': str(row.get('is_compliant', 'Unknown')),
                'gap_description': str(row.get('gap_description', '')) if not pd.isna(row.get('gap_description', '')) else '',
                'gap_recommendations': str(row.get('gap_recommendations', '')) if not pd.isna(row.get('gap_recommendations', '')) else '',
                'applicability': str(row.get('applicability', 'Unknown'))
            }
            detailed_results.append(result)
        
        document_data = {
            'document_name': str(document_name),
            'detailed_results': detailed_results
        }
        grouped_data.append(document_data)
    
    return grouped_data

def load_regulation_metadata():
    """Load the regulation metadata JSON file."""
    base_dir = Path(r"C:\Users\91810\OneDrive\Desktop\CPC_AIB_Life")
    json_file_path = base_dir / "output" / "Data_extraction" / "regulation_17A_48_combined.json"
    
    print(f"Loading regulation metadata from {json_file_path}...")
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Create a lookup dictionary for easy access
    regulation_lookup = {}
    for regulation in data:
        key = (regulation.get('Section Type', ''), str(regulation.get('Regulation Number', '')))
        regulation_lookup[key] = regulation
    
    return regulation_lookup

def group_by_regulation(compliance_data):
    """Group compliance results by regulation."""
    grouped_results = {}
    
    print("Grouping compliance results by regulation...")
    for document_analysis in compliance_data:
        document_name = document_analysis.get('document_name', 'Unknown Document')
        detailed_results = document_analysis.get('detailed_results', [])
        
        for result in detailed_results:
            section_type = result.get('section_type', '')
            regulation_number = result.get('regulation_number', '')
            
            # Create a unique key for each regulation
            reg_key = (section_type, str(regulation_number))
            
            if reg_key not in grouped_results:
                grouped_results[reg_key] = {
                    'section_type': section_type,
                    'source_name': result.get('source_name', ''),
                    'regulation_number': regulation_number,
                    'regulation_title': result.get('regulation_title', ''),
                    'regulation_text': result.get('regulation_text', ''),
                    'documents': [],
                    'compliance_statuses': [],
                    'gap_descriptions': [],
                    'gap_recommendations': [],
                    'applicability': []  # Track applicability
                }
            
            # Add document-specific information
            grouped_results[reg_key]['documents'].append(document_name)
            grouped_results[reg_key]['compliance_statuses'].append(result.get('is_compliant', 'Unknown'))
            grouped_results[reg_key]['applicability'].append(result.get('applicability', 'Unknown'))
            
            # Add gaps and recommendations if available (only for applicable regulations)
            applicability = result.get('applicability', '')
            if applicability in ['Applies', 'May Apply - Requires Further Review']:
                if result.get('gap_description'):
                    grouped_results[reg_key]['gap_descriptions'].append(result.get('gap_description'))
                
                if result.get('gap_recommendations'):
                    grouped_results[reg_key]['gap_recommendations'].append(result.get('gap_recommendations'))
    
    return grouped_results

def count_compliance_status(regulation_data):
    """Count compliance statuses for a regulation.
    Only count documents where applicability is 'Applies' or 'May Apply - Requires Further Review'
    """
    # Create a list of tuples with (document, status, applicability) for easier filtering
    document_status_list = list(zip(
        regulation_data['documents'], 
        regulation_data['compliance_statuses'],
        regulation_data['applicability']
    ))
    
    # Filter to only include applicable documents
    applicable_docs = [
        (doc, status, app) for doc, status, app in document_status_list 
        if app in ['Applies', 'May Apply - Requires Further Review']
    ]
    
    # Count unique applicable documents
    unique_applicable_docs = set(doc for doc, _, _ in applicable_docs)
    documents_verified_count = len(unique_applicable_docs)
    
    # Count compliance statuses only for applicable documents
    yes_count = sum(1 for _, status, _ in applicable_docs if status == 'Yes')
    partial_count = sum(1 for _, status, _ in applicable_docs if status == 'Partial')
    no_count = sum(1 for _, status, _ in applicable_docs if status == 'No')
    
    return {
        'documents_verified_count': documents_verified_count,
        'is_compliant_yes_count': yes_count,
        'is_compliant_partial_count': partial_count,
        'is_compliant_no_count': no_count
    }

def determine_priority(compliance_counts, gap_summary):
    """Determine the priority level (High, Medium, Low) based on compliance status and gaps."""
    # Extract counts
    yes_count = compliance_counts['is_compliant_yes_count']
    partial_count = compliance_counts['is_compliant_partial_count']
    no_count = compliance_counts['is_compliant_no_count']
    total_count = compliance_counts['documents_verified_count']
    
    # Calculate compliance percentage
    if total_count == 0:
        return "Low"  # No documents verified
    
    compliance_percentage = (yes_count + (partial_count * 0.5)) / total_count
    
    # Check for critical keywords in gap summary
    critical_keywords = [
        "risk", "critical", "severe", "urgent", "regulatory", "penalty", 
        "fine", "breach", "violation", "customer harm", "vulnerable"
    ]
    
    has_critical_gaps = any(keyword in gap_summary.lower() for keyword in critical_keywords) if gap_summary else False
    
    # Determine priority
    if no_count > 0 and no_count / total_count > 0.3:
        # More than 30% non-compliant
        return "High"
    elif has_critical_gaps or compliance_percentage < 0.7:
        # Contains critical gaps or less than 70% compliant
        return "Medium"
    else:
        # Mostly compliant without critical gaps
        return "Low"

def generate_detailed_bullet_points(gaps, recommendations, regulation_info):
    """Generate detailed bullet points for gaps and recommendations using AI API.
    The number of bullet points is dynamic based on the number of applicable documents."""
    # Check if this regulation is not applicable
    applicability_list = regulation_info.get('applicability', [])
    if applicability_list and all(app == "Does Not Apply" for app in applicability_list if app):
        return "• **Not Applicable:** This regulation does not apply to the reviewed documents.", ""
    
    if not gaps and not recommendations:
        return "", ""
    
    # Calculate the number of applicable documents
    applicable_docs = sum(1 for app in applicability_list if app in ['Applies', 'May Apply - Requires Further Review'])
    
    # Determine the number of bullet points based on the number of applicable documents
    if applicable_docs <= 5:
        num_points = "5-7"  # Few documents, keep original range
    elif applicable_docs <= 20:
        num_points = "7-9"  # Medium number of documents
    else:
        num_points = "10-12"  # Many documents (close to or more than 20)
    
    # Prepare the prompt for gap descriptions
    gap_prompt = f"""
    You are analyzing compliance gaps for financial regulations. Below are multiple gap descriptions 
    identified across different documents for the same regulation:
    
    Regulation: {regulation_info.get('regulation_title', 'Unknown')}
    Regulation Text: {regulation_info.get('regulation_text', 'Unknown')}
    Number of Documents Analyzed: {applicable_docs}
    
    Gap Descriptions:
    {' '.join(gaps)}
    
    Please summarize these gaps into {num_points} key points. Format each point as a detailed bullet point with a bold header 
    followed by a specific description. Focus on the most critical issues first. 
    
    Example format:
    • **Header for Issue 1:** Detailed description of the issue with specific context.
    • **Header for Issue 2:** Detailed description of the issue with specific context.
    
    Only return the bullet points, nothing else.
    """
    
    # Prepare the prompt for recommendations
    rec_prompt = f"""
    You are providing recommendations to address compliance gaps for financial regulations. Below are multiple 
    recommendations identified across different documents for the same regulation:
    
    Regulation: {regulation_info.get('regulation_title', 'Unknown')}
    Regulation Text: {regulation_info.get('regulation_text', 'Unknown')}
    Number of Documents Analyzed: {applicable_docs}
    
    Recommendations:
    {' '.join(recommendations)}
    
    Please summarize these recommendations into {num_points} key actionable points. Format each point as a detailed bullet point 
    with a bold header followed by a specific action plan. Focus on the most critical actions first.
    
    Example format:
    • **Header for Action 1:** Detailed description of the recommended action with specific implementation guidance.
    • **Header for Action 2:** Detailed description of the recommended action with specific implementation guidance.
    
    Only return the bullet points, nothing else.
    """
    
    # Call AI API for gap descriptions
    gap_summary = call_ai_api(gap_prompt) if gaps else ""
    
    # Call AI API for recommendations
    rec_summary = call_ai_api(rec_prompt) if recommendations else ""
    
    return gap_summary, rec_summary

def call_ai_api(prompt):
    """Call the GPT-4o API via OpenRouter."""
    if not API_KEY:
        print("Warning: API_KEY not set. Using mock response.")
        return "• **Mock Response:** This is a placeholder. Set API_KEY to get actual AI-generated summaries."
    
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    
    data = {
        "model": "openai/gpt-4o",  # Using GPT-4o instead of Claude
        "messages": [
            {"role": "user", "content": prompt}
        ],
        "max_tokens": 1000
    }
    
    try:
        response = requests.post(API_URL, headers=headers, json=data)
        response.raise_for_status()
        result = response.json()
        return result['choices'][0]['message']['content']
    except Exception as e:
        print(f"Error calling Claude API: {e}")
        return "Error generating summary"

def process_regulation_group(reg_key, reg_data, regulation_metadata):
    """Process a single regulation group with API calls."""
    print(f"Processing regulation {reg_key}...")
    
    # Get counts
    counts = count_compliance_status(reg_data)
    
    # Get metadata
    metadata = regulation_metadata.get(reg_key, {})
    
    # Use regulation text from metadata if available
    regulation_text = metadata.get('Regulation Text', '') if metadata and metadata.get('Regulation Text') else reg_data['regulation_text']
    regulation_title = metadata.get('Regulation Title', '') if metadata and metadata.get('Regulation Title') else reg_data['regulation_title']
    
    # Generate summaries
    gap_summary, rec_summary = generate_detailed_bullet_points(
        reg_data['gap_descriptions'], 
        reg_data['gap_recommendations'],
        {
            'regulation_title': regulation_title,
            'regulation_text': regulation_text,
            'applicability': reg_data.get('applicability', [])
        }
    )
    
    # Determine priority
    priority = determine_priority(counts, gap_summary)
    
    # Create consolidated record
    consolidated = {
        'section_type': reg_data['section_type'],
        'source_name': reg_data['source_name'],
        'part_number': metadata.get('Part Number', ''),
        'part_name': metadata.get('Part Name', ''),
        'chapter_number': metadata.get('Chapter Number', ''),
        'chapter_name': metadata.get('Chapter Name', ''),
        'regulation_number': reg_data['regulation_number'],
        'regulation_title': regulation_title,
        'regulation_text': regulation_text,
        'priority': priority,  # Add priority
        'documents_verified_count': counts['documents_verified_count'],
        'is_compliant_yes_count': counts['is_compliant_yes_count'],
        'is_compliant_partial_count': counts['is_compliant_partial_count'],
        'is_compliant_no_count': counts['is_compliant_no_count'],
        'gap_description': gap_summary,
        'gap_recommendations': rec_summary
    }
    
    return consolidated

def create_consolidated_excel(consolidated_data, output_path):
    """Create a consolidated Excel report with priority highlighting."""
    print(f"Creating Excel report at {output_path}...")
    
    # Convert to DataFrame
    df = pd.DataFrame(consolidated_data)
    
    # Sort by priority first, then section_type, part_number, etc.
    priority_order = {'High': 0, 'Medium': 1, 'Low': 2}
    df['priority_sort'] = df['priority'].map(priority_order)
    df['part_number_sort'] = pd.to_numeric(df['part_number'], errors='coerce').fillna(9999)
    df['chapter_number_sort'] = pd.to_numeric(df['chapter_number'], errors='coerce').fillna(9999)
    df['regulation_number_sort'] = pd.to_numeric(df['regulation_number'], errors='coerce').fillna(9999)
    
    df = df.sort_values(by=['priority_sort', 'section_type', 'part_number_sort', 'chapter_number_sort', 'regulation_number_sort'])
    
    # Drop sorting columns
    df = df.drop(columns=['priority_sort', 'part_number_sort', 'chapter_number_sort', 'regulation_number_sort'])
    
    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write main compliance data
        df.to_excel(writer, sheet_name='Consolidated Compliance', index=False)
        
        # Create definitions sheet
        create_definitions_sheet(writer)
        
        # Get the workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Consolidated Compliance']
        
        # Define styles for priorities
        red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
        green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        
        bold_font = Font(bold=True)
        
        # Find the priority column index
        priority_col_idx = None
        for idx, col in enumerate(df.columns):
            if col == 'priority':
                priority_col_idx = idx + 1  # +1 because Excel is 1-indexed
                break
        
        # Apply conditional formatting based on priority
        if priority_col_idx:
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=priority_col_idx, max_col=priority_col_idx)):
                cell = row[0]
                if cell.value == 'High':
                    cell.fill = red_fill
                    cell.font = bold_font
                elif cell.value == 'Medium':
                    cell.fill = yellow_fill
                elif cell.value == 'Low':
                    cell.fill = green_fill
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Limit maximum column width
            adjusted_width = min(max_length + 2, 100)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add filters
        worksheet.auto_filter.ref = worksheet.dimensions
    
    print(f"Excel report created successfully at {output_path}")

def create_definitions_sheet(excel_writer):
    """Create a definitions sheet explaining each column in the report."""
    # Define column definitions
    definitions = [
        ["Column Name", "Definition"],
        ["section_type", "The type of regulatory section (e.g., '17A' for Central Bank Reform Act 2010 Section 17A, or '48' for Central Bank Supervision and Enforcement Act 2013 Section 48)"],
        ["source_name", "The full name of the regulatory source document"],
        ["part_number", "The part number within the regulation"],
        ["part_name", "The name of the part within the regulation"],
        ["chapter_number", "The chapter number within the part (if applicable)"],
        ["chapter_name", "The name of the chapter within the part (if applicable)"],
        ["regulation_number", "The specific regulation number"],
        ["regulation_title", "The title of the regulation"],
        ["regulation_text", "The full text of the regulation requirement"],
        ["priority", "The assigned priority level (High, Medium, Low) based on compliance status and critical keywords in gaps"],
        ["documents_verified_count", "The number of unique documents where this regulation was applicable (either 'Applies' or 'May Apply - Requires Further Review')"],
        ["is_compliant_yes_count", "The number of documents that were fully compliant with this regulation"],
        ["is_compliant_partial_count", "The number of documents that were partially compliant with this regulation"],
        ["is_compliant_no_count", "The number of documents that were not compliant with this regulation"],
        ["gap_description", "AI-generated summary of compliance gaps identified across all applicable documents"],
        ["gap_recommendations", "AI-generated recommendations to address the identified compliance gaps"]
    ]
    
    # Create DataFrame for definitions
    df_definitions = pd.DataFrame(definitions[1:], columns=definitions[0])
    
    # Write to Excel
    df_definitions.to_excel(excel_writer, sheet_name='Definitions', index=False)
    
    # Format the definitions sheet
    worksheet = excel_writer.sheets['Definitions']
    
    # Apply formatting
    bold_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = bold_font
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set column width
        adjusted_width = min(max_length + 2, 100)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def generate_consolidated_report():
    """Generate a consolidated compliance report."""
    # Load data
    compliance_data = load_compliance_results()
    regulation_metadata = load_regulation_metadata()
    
    # Group by regulation
    grouped_results = group_by_regulation(compliance_data)
    
    print(f"Processing {len(grouped_results)} unique regulations...")
    
    # Process each regulation group
    consolidated_results = []
    
    # Use ThreadPoolExecutor for parallel processing
    with ThreadPoolExecutor(max_workers=5) as executor:
        # Submit all tasks
        future_to_reg = {
            executor.submit(
                process_regulation_group, reg_key, reg_data, regulation_metadata
            ): reg_key 
            for reg_key, reg_data in grouped_results.items()
        }
        
        # Process results as they complete
        for future in future_to_reg:
            reg_key = future_to_reg[future]
            try:
                result = future.result()
                consolidated_results.append(result)
                print(f"Completed processing regulation {reg_key}")
            except Exception as e:
                print(f"Error processing regulation {reg_key}: {e}")
    
    # Create output directory if it doesn't exist
    base_dir = Path(r"C:\Users\91810\OneDrive\Desktop\CPC_AIB_Life")
    output_dir = base_dir / "output" / "enhanced_document_analysis"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Save consolidated results to JSON
    json_output_path = output_dir / "consolidated_compliance_report.json"
    with open(json_output_path, 'w', encoding='utf-8') as f:
        json.dump(consolidated_results, f, indent=2)
    
    print(f"Consolidated JSON saved to {json_output_path}")
    
    # Create Excel report
    excel_output_path = output_dir / "consolidated_compliance_report.xlsx"
    create_consolidated_excel(consolidated_results, excel_output_path)

if __name__ == "__main__":
    generate_consolidated_report()
